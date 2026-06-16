"""i18n.js ↔ Excel 互转工具
用法:
  python i18n_export.py            → i18n.js → i18n.xlsx
  python i18n_export.py --import   → i18n.xlsx → i18n.js
"""
import subprocess, json, sys, os

HERE = os.path.dirname(os.path.abspath(__file__))
JS_FILE = os.path.join(HERE, 'i18n.js')
XLSX_FILE = os.path.join(HERE, '..', 'i18n.xlsx')

def dump_json():
    """用 Node 加载 i18n.js，输出 zh/en 为 JSON"""
    script = """
var fs = require('fs');
var vm = require('vm');
var code = fs.readFileSync(process.argv[1], 'utf8');
var ctx = { window: {} };
vm.createContext(ctx);
vm.runInContext(code, ctx);
var data = ctx.window.I18N;
console.log(JSON.stringify(data));
"""
    result = subprocess.run(['node', '-e', script, JS_FILE], capture_output=True, text=True, encoding='utf-8')
    if result.returncode != 0:
        raise RuntimeError(f"Node error: {result.stderr}")
    return json.loads(result.stdout)

def export_xlsx():
    data = dump_json()
    zh = data['zh']
    en = data['en']
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'i18n'
    ws.append(['key', '中文', '英文'])
    for key in sorted(zh.keys()):
        ws.append([key, zh.get(key, ''), en.get(key, '')])
    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 65
    ws.column_dimensions['C'].width = 65
    wb.save(XLSX_FILE)
    print(f'已导出 {len(zh)} 条 → {XLSX_FILE}')

def import_xlsx():
    import openpyxl
    wb = openpyxl.load_workbook(XLSX_FILE)
    ws = wb.active
    zh_entries = []
    en_entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        key, zh_val, en_val = row[0], row[1] or '', row[2] or ''
        if key:
            zh_entries.append(f'    "{key}": {json.dumps(zh_val, ensure_ascii=False)}')
            en_entries.append(f'    "{key}": {json.dumps(en_val, ensure_ascii=False)}')

    output = f'''// 厢间人 UI 文案数据库 — 中英双语
// 用 i18n_export.py 导出 Excel 修改，改完用 i18n_export.py --import 导回
window.I18N = {{
  zh: {{
{",\n".join(zh_entries)}
  }},
  en: {{
{",\n".join(en_entries)}
  }}
}};

function T(key) {{
  var d = window.I18N[STATE.lang];
  return (d && d[key]) ? d[key] : key;
}}
'''
    with open(JS_FILE, 'w', encoding='utf-8') as f:
        f.write(output)
    print(f'已导入 → {JS_FILE}')

if __name__ == '__main__':
    if '--import' in sys.argv:
        import_xlsx()
    else:
        export_xlsx()
