"""quiz.js ↔ Excel 互转工具
用法:
  python quiz_export.py            → quiz.js → quiz.xlsx
  python quiz_export.py --import   → quiz.xlsx → quiz.js
"""
import subprocess, json, sys, os

HERE = os.path.dirname(os.path.abspath(__file__))
JS_FILE = os.path.join(HERE, 'quiz.js')
XLSX_FILE = os.path.join(HERE, '..', 'quiz.xlsx')

def dump_json():
    script = """
var fs = require('fs'); var vm = require('vm');
var code = fs.readFileSync(process.argv[1], 'utf8');
var ctx = { window: {} }; vm.createContext(ctx); vm.runInContext(code, ctx);
console.log(JSON.stringify({ zh: ctx.window.QUIZ_ZH, en: ctx.window.QUIZ_EN }));
"""
    result = subprocess.run(['node', '-e', script, JS_FILE], capture_output=True, text=True, encoding='utf-8')
    if result.returncode != 0: raise RuntimeError(f"Node error: {result.stderr}")
    return json.loads(result.stdout)

SKILL_LABEL = ["哲学思辨/Phil.Thought", "逻辑推导/Logical Deduction", "缪斯女神/The Muse", "灵魂叙事/Soul Narrative", "技术专家/Technical Expert", "身强体壮/Physical Might"]

def export_xlsx():
    data = dump_json()
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '问卷'
    ws.append(['题号', '语言', '题目', '选项A', 'A权重', '选项B', 'B权重', '选项C', 'C权重'])
    for lang in ['zh', 'en']:
        for idx, q in enumerate(data[lang]):
            row = [idx + 1, lang, q['q']]
            for oi in range(3):
                row.append(q['a'][oi])
                s = q['s'][oi]
                parts = []
                for k, v in s.items():
                    parts.append(f"{SKILL_LABEL[int(k)]}={v}")
                row.append("; ".join(parts))
            ws.append(row)
    for col, w in enumerate([8, 6, 50, 40, 30, 40, 30, 40, 30], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    wb.save(XLSX_FILE)
    print(f'已导出 {len(data["zh"])*2} 条 → {XLSX_FILE}')

def import_xlsx():
    import openpyxl, re
    wb = openpyxl.load_workbook(XLSX_FILE)
    ws = wb.active
    zh_data, en_data = [], []
    name_to_idx = {"哲学思辨":0,"逻辑推导":1,"缪斯女神":2,"灵魂叙事":3,"技术专家":4,"身强体壮":5,
                   "Phil.Thought":0,"Logical Deduction":1,"The Muse":2,"Soul Narrative":3,"Technical Expert":4,"Physical Might":5}
    for row in ws.iter_rows(min_row=2, values_only=True):
        num, lang, q = row[0], row[1], row[2]
        if not q: continue
        entry = {"q": q, "a": [], "s": []}
        for oi in range(3):
            entry["a"].append(row[3 + oi*2] or "")
            weight_str = row[4 + oi*2] or ""
            s = {}
            for part in weight_str.split(";"):
                part = part.strip()
                if "=" in part:
                    name, val = part.rsplit("=", 1)
                    name = name.strip()
                    for k, v in name_to_idx.items():
                        if k in name: s[str(v)] = int(val.strip()); break
            entry["s"].append(s)
        (zh_data if lang == 'zh' else en_data).append(entry)

    zh_entries = json.dumps(zh_data, ensure_ascii=False, indent=2)
    en_entries = json.dumps(en_data, ensure_ascii=False, indent=2)
    output = f'''// 厢间人 问卷数据库 — 中英双语
// 用 quiz_export.py 导出 Excel 修改，改完用 quiz_export.py --import 导回
// 权重格式: {{技能索引: 分值}}，索引 0=哲学思辨 1=逻辑推导 2=缪斯女神 3=灵魂叙事 4=技术专家 5=身强体壮
window.QUIZ_ZH = {zh_entries};
window.QUIZ_EN = {en_entries};
function QUIZ() {{ return STATE.lang === "zh" ? window.QUIZ_ZH : window.QUIZ_EN; }}
'''
    with open(JS_FILE, 'w', encoding='utf-8') as f: f.write(output)
    print(f'已导入 → {JS_FILE}')

if __name__ == '__main__':
    if '--import' in sys.argv: import_xlsx()
    else: export_xlsx()
